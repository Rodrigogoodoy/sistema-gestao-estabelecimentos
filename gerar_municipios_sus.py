"""
Gera MUNICIPIOS_SUS_SIM.xlsx a partir de resultados_municipios_com_cnpj.xlsx.
Todos os registros já atendem SUS (filtro aplicado na coleta).
"""
import datetime
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ENTRADA = "resultados_fase2.xlsx"
SAIDA   = "MUNICIPIOS_SUS_SIM.xlsx"

COLUNAS_MAP = {
    "CNPJ":                          "CNPJ",
    "CNPJ Mantenedora":              "CNPJ Mantenedora",
    "Município":                     "Municipio",
    "UF":                            "UF",
    "CNES":                          "Codigo CNES",
    "Nome Fantasia":                 "Nome Fantasia",
    "Nome":                          "Razao Social",
    "Tipo de Estabelecimento":       "Tipo de Estabelecimento",
    "Classificação Estabelecimento": "Classificacao",
    "Gestão":                        "Gestao",
    "Natureza Jurídica(Grupo)":      "Natureza Juridica",
    "CEP":                           "CEP",
    "Logradouro":                    "Logradouro",
    "Número":                        "Numero",
    "Bairro":                        "Bairro",
    "Complemento":                   "Complemento",
    "Telefone":                      "Telefone",
    "E-mail":                        "Email",
    "Responsável Técnico":           "Responsavel Tecnico",
    "Data Desativação":              "Data Desativacao",
    "Motivo Desativação":            "Motivo Desativacao",
}

LARGURAS = {
    "CNPJ": 20, "Municipio": 28, "UF": 8,
    "Codigo CNES": 14, "Nome Fantasia": 35, "Razao Social": 38,
    "Tipo de Estabelecimento": 30, "Classificacao": 30, "Gestao": 16,
    "Natureza Juridica": 30, "CEP": 14, "Logradouro": 32,
    "Numero": 10, "Bairro": 20, "Complemento": 20,
    "CNPJ Mantenedora": 22,
    "Telefone": 18, "Email": 28, "Responsavel Tecnico": 30,
    "Data Desativacao": 18, "Motivo Desativacao": 40,
}


def gerar():
    df = pd.read_excel(ENTRADA, dtype=str)
    print(f"Registros carregados: {len(df)}")

    # Mantém apenas colunas mapeadas que existem no arquivo
    cols_presentes = {k: v for k, v in COLUNAS_MAP.items() if k in df.columns}
    df_exp = df[list(cols_presentes.keys())].rename(columns=cols_presentes)

    # Garante colunas de desativação mesmo se ausentes
    for col in ("Data Desativacao", "Motivo Desativacao"):
        if col not in df_exp.columns:
            df_exp[col] = ""

    # Limpa valores inválidos
    df_exp = df_exp.replace({"null - null": "", "null": ""})
    df_exp = df_exp.fillna("")

    qtd = len(df_exp)
    agora = datetime.datetime.now().strftime("%d/%m/%Y")

    writer = pd.ExcelWriter(SAIDA, engine="openpyxl")
    df_exp.to_excel(writer, sheet_name="Estabelecimentos", index=False, startrow=4)
    ws = writer.sheets["Estabelecimentos"]

    def fill(cor): return PatternFill("solid", fgColor=cor)
    def borda():
        lado = Side(style="thin", color="D1D5DB")
        return Border(left=lado, right=lado, top=lado, bottom=lado)

    ncols = df_exp.shape[1]
    ult = get_column_letter(ncols)

    # Linha 1 — título
    ws.merge_cells(f"A1:{ult}1")
    ws["A1"] = "ESTABELECIMENTOS DE SAUDE — ATENDE SUS: SIM (BUSCA POR MUNICÍPIO)"
    ws["A1"].fill = fill("1E3A8A")
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    # Linha 2 — subtítulo
    ws.merge_cells(f"A2:{ult}2")
    ws["A2"] = (
        f"Data de geracao: {agora}"
        f"   |   Estabelecimentos: {qtd:,}"
        f"   |   Filtro: Atende SUS = Sim"
        f"   |   Fonte: cnes.datasus.gov.br"
    )
    ws["A2"].fill = fill("2563EB")
    ws["A2"].font = Font(bold=False, color="FFFFFF", size=10, name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # Linha 3 — nota
    ativos = int((df_exp["Data Desativacao"] == "").sum())
    inativos = qtd - ativos
    ws.merge_cells(f"A3:{ult}3")
    ws["A3"] = f"Ativos: {ativos:,}  |  Com data de desativação: {inativos:,}"
    ws["A3"].fill = fill("DBEAFE")
    ws["A3"].font = Font(bold=False, color="1E3A8A", size=10, name="Calibri")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20

    ws.row_dimensions[4].height = 8

    # Cabeçalho
    for col in range(1, ncols + 1):
        c = ws.cell(row=5, column=col)
        c.fill = fill("1E3A8A")
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda()
    ws.row_dimensions[5].height = 30

    # Dados
    for row in range(6, qtd + 6):
        bg = "F0F9FF" if row % 2 == 0 else "FFFFFF"
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = fill(bg)
            c.font = Font(bold=False, color="1E293B", size=10, name="Calibri")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = borda()
        ws.row_dimensions[row].height = 18

    # Larguras
    for i, col_name in enumerate(df_exp.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = LARGURAS.get(col_name, 18)

    ws.freeze_panes = "A6"
    writer.close()
    print(f"Gerado: {SAIDA} ({qtd} registros | {ativos} ativos | {inativos} com desativação)")


if __name__ == "__main__":
    gerar()
