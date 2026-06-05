import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

workers = [
    'resultados_worker_0.xlsx',
    'resultados_worker_1.xlsx',
    'resultados_worker_2.xlsx',
    'resultados_worker_3.xlsx',
]

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "SIM"

header_written = False
total_sim = 0

for fname in workers:
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not header_written:
        ws_out.append(list(rows[0]))
        header_written = True
    for row in rows[1:]:
        if row[1] is not None and str(row[1]).strip().upper() == 'SIM':
            ws_out.append(list(row))
            total_sim += 1

# Estilo do cabeçalho
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
for cell in ws_out[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Largura automática das colunas (limitada a 50)
for col_idx in range(1, ws_out.max_column + 1):
    max_len = 0
    col_letter = get_column_letter(col_idx)
    for row in ws_out.iter_rows(min_col=col_idx, max_col=col_idx):
        val = str(row[0].value) if row[0].value is not None else ""
        if len(val) > max_len:
            max_len = len(val)
    ws_out.column_dimensions[col_letter].width = min(max_len + 2, 50)

# Congelar primeira linha
ws_out.freeze_panes = "A2"

# Filtro automático
ws_out.auto_filter.ref = ws_out.dimensions

output = 'resultados_SIM.xlsx'
wb_out.save(output)
print(f"Arquivo salvo: {output}")
print(f"Total de linhas com SIM: {total_sim}")
