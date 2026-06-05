"""
Re-busca os CNPJs já encontrados (SIM) com o filtro "Atende SUS = Sim" ativo.
Se o CNPJ aparecer na busca com filtro = atende SUS. Se não aparecer = não atende.
Não precisa entrar na ficha — mais rápido.

Uso:
  python recheck_sus.py                          # 1 worker, todos os registros
  python recheck_sus.py --worker 0 --workers-total 4
  python recheck_sus.py --merge                  # mescla e gera ESTABELECIMENTOS_SUS_SIM.xlsx
"""
import argparse
import json
import time
import pandas as pd
from pathlib import Path
from tqdm import tqdm
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

from config import TARGET_URL, HEADLESS

TIMEOUT = 20_000

_JS_PEGAR_HREF = (
    "() => {"
    "  const links = [...document.querySelectorAll('a[href*=\"coUnidade=\"]')]"
    "    .filter(a => { const m = (a.getAttribute('href')||'').match(/coUnidade=(\\d+)/); return m && parseInt(m[1]) > 0; });"
    "  return links.length > 0 ? links[0].href : null;"
    "}"
)


def _ativar_filtro_sus(page):
    """Clica no botão 'Sim' do filtro Atende SUS na página de busca."""
    loc = page.locator("button:has-text('Sim')")
    for i in range(loc.count()):
        el = loc.nth(i)
        if el.is_visible() and el.inner_text().strip() == "Sim":
            el.click()
            time.sleep(0.3)
            return True
    return False


def _buscar_cnpj_com_sus(page, cnpj: str) -> bool:
    """
    Abre a página de busca, ativa SUS=Sim, pesquisa o CNPJ.
    Retorna True se apareceu resultado (= atende SUS).
    """
    page.goto(TARGET_URL, timeout=TIMEOUT)
    page.wait_for_load_state("networkidle", timeout=TIMEOUT)

    _ativar_filtro_sus(page)

    seletores = [
        "input[ng-model*='cnpj']", "input[ng-model*='Cnpj']",
        "input[ng-model*='nuCnpj']", "input[placeholder*='CNPJ']",
        "input[name*='cnpj']", "#cnpj",
    ]
    campo = None
    for sel in seletores:
        loc = page.locator(sel)
        if loc.count() > 0:
            campo = loc.first
            break
    if not campo:
        return False

    campo.fill(cnpj)
    time.sleep(0.3)
    campo.press("Tab")
    time.sleep(0.3)

    page.locator("button:has-text('Pesquisar')").first.click()
    page.wait_for_timeout(700)
    page.wait_for_load_state("networkidle", timeout=TIMEOUT)
    page.wait_for_timeout(400)

    href = page.evaluate(_JS_PEGAR_HREF)
    return href is not None


def processar_worker(worker_id: int, workers_total: int):
    entrada = Path("ESTABELECIMENTOS_CNES.xlsx")
    if not entrada.exists():
        print("ESTABELECIMENTOS_CNES.xlsx não encontrado. Rode merge_results.py primeiro.")
        return

    df = pd.read_excel(entrada, skiprows=4, dtype=str)
    total = len(df)
    print(f"Total de registros SIM: {total}")

    tamanho = (total + workers_total - 1) // workers_total
    inicio_slice = worker_id * tamanho
    fim_slice = min(inicio_slice + tamanho, total)
    fatia = df.iloc[inicio_slice:fim_slice].copy().reset_index(drop=True)

    checkpoint_file = Path(f"checkpoint_sus_worker_{worker_id}.json")

    inicio = 0
    sus_map = {}
    if checkpoint_file.exists():
        try:
            data = json.loads(checkpoint_file.read_text(encoding="utf-8"))
            inicio = data.get("indice", 0)
            sus_map = data.get("sus_map", {})
            print(f"[worker {worker_id}] Retomando do índice {inicio} ({len(sus_map)} já verificados)")
        except Exception:
            pass

    pendentes = len(fatia) - inicio
    print(f"[worker {worker_id}] Registros a verificar: {pendentes} ({inicio_slice+1}–{fim_slice} de {total})\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        page = browser.new_page()

        with tqdm(total=pendentes, desc=f"Worker {worker_id}", unit="cnpj") as barra:
            for i in range(inicio, len(fatia)):
                cnpj_raw = str(fatia.iloc[i].get("CNPJ", "")).strip()
                cnpj = cnpj_raw.replace(".", "").replace("/", "").replace("-", "")

                sus_val = "ERRO"
                for tentativa in range(1, 3):
                    try:
                        atende = _buscar_cnpj_com_sus(page, cnpj)
                        sus_val = "Sim" if atende else "Não"
                        break
                    except PlaywrightTimeout:
                        if tentativa == 2:
                            sus_val = "ERRO"
                        time.sleep(3)
                    except Exception:
                        if tentativa == 2:
                            sus_val = "ERRO"
                        time.sleep(2)

                sus_map[cnpj_raw] = sus_val
                checkpoint_file.write_text(
                    json.dumps({"indice": i + 1, "sus_map": sus_map}),
                    encoding="utf-8"
                )
                barra.update(1)
                sim = sum(1 for v in sus_map.values() if v == "Sim")
                nao = sum(1 for v in sus_map.values() if v == "Não")
                barra.set_postfix(SIM=sim, NAO=nao)

        browser.close()

    # Salva parcial deste worker
    output_file = Path(f"sus_worker_{worker_id}.xlsx")
    fatia["Atende SUS"] = fatia["CNPJ"].map(sus_map).fillna("ERRO")
    fatia.to_excel(output_file, index=False)
    sim = int((fatia["Atende SUS"] == "Sim").sum())
    print(f"\n[worker {worker_id}] Concluído! SIM={sim} | Salvo em {output_file}")

    if checkpoint_file.exists():
        checkpoint_file.unlink()


def mesclar_e_gerar():
    import datetime
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    arquivos = sorted(Path(".").glob("sus_worker_*.xlsx"))
    if not arquivos:
        print("Nenhum sus_worker_*.xlsx encontrado.")
        return

    dfs = [pd.read_excel(a, dtype=str) for a in arquivos]
    df = pd.concat(dfs, ignore_index=True)

    total = len(df)
    sus_sim = int((df.get("Atende SUS", pd.Series(dtype=str)) == "Sim").sum())
    print(f"Total SIM no CNES: {total:,} | Atende SUS=Sim: {sus_sim:,}")

    df_sim = df[df["Atende SUS"] == "Sim"].copy().reset_index(drop=True)
    df_sim = df_sim.replace({"null - null": "", "null": ""})

    # Ordem das colunas: CNPJ, Municipio, Atende SUS, demais
    cols = list(df_sim.columns)
    prioridade = ["CNPJ", "Municipio", "Atende SUS"]
    ordenadas = [c for c in prioridade if c in cols] + [c for c in cols if c not in prioridade]
    df_sim = df_sim[ordenadas]

    agora = datetime.datetime.now().strftime("%d/%m/%Y")
    output = "ESTABELECIMENTOS_SUS_SIM.xlsx"

    writer = pd.ExcelWriter(output, engine="openpyxl")
    df_sim.to_excel(writer, sheet_name="Estabelecimentos", index=False, startrow=4)
    ws = writer.sheets["Estabelecimentos"]

    def fill(cor): return PatternFill("solid", fgColor=cor)
    def borda():
        lado = Side(style="thin", color="D1D5DB")
        return Border(left=lado, right=lado, top=lado, bottom=lado)

    ncols = df_sim.shape[1]
    ult = get_column_letter(ncols)

    ws.merge_cells(f"A1:{ult}1")
    ws["A1"] = "ESTABELECIMENTOS DE SAUDE — ATENDE SUS: SIM"
    ws["A1"].fill = fill("1E3A8A")
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells(f"A2:{ult}2")
    ws["A2"] = (
        f"Data de geracao: {agora}"
        f"   |   Total SIM no CNES: {total:,}"
        f"   |   Atende SUS: {sus_sim:,}"
        f"   |   Fonte: cnes.datasus.gov.br"
    )
    ws["A2"].fill = fill("2563EB")
    ws["A2"].font = Font(bold=False, color="FFFFFF", size=10, name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    ws.merge_cells(f"A3:{ult}3")
    ws["A3"] = f"Lista com {sus_sim:,} estabelecimentos confirmados no CNES que atendem o SUS"
    ws["A3"].fill = fill("DBEAFE")
    ws["A3"].font = Font(bold=False, color="1E3A8A", size=10, name="Calibri")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 8

    for col in range(1, ncols + 1):
        c = ws.cell(row=5, column=col)
        c.fill = fill("1E3A8A")
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda()
    ws.row_dimensions[5].height = 30

    for row in range(6, sus_sim + 6):
        bg = "F0F9FF" if row % 2 == 0 else "FFFFFF"
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = fill(bg)
            c.font = Font(bold=False, color="1E293B", size=10, name="Calibri")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = borda()
        ws.row_dimensions[row].height = 18

    larguras = {
        "CNPJ": 20, "Municipio": 28, "Atende SUS": 14,
        "Codigo CNES": 14, "Nome Fantasia": 35, "Razao Social": 38,
        "Tipo de Estabelecimento": 30, "Classificacao": 30, "Gestao": 16,
        "Natureza Juridica": 30, "UF": 8, "CEP": 14,
        "Logradouro": 32, "Numero": 10, "Bairro": 20,
        "Telefone": 18, "Email": 28, "Responsavel Tecnico": 30,
        "Data Desativacao": 18, "Motivo Desativacao": 35,
    }
    for i, col_name in enumerate(df_sim.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(col_name, 18)

    ws.freeze_panes = "A6"
    writer.close()
    print(f"Arquivo gerado: {output} ({sus_sim:,} registros)")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--worker", type=int, default=0)
    parser.add_argument("--workers-total", type=int, default=1)
    parser.add_argument("--merge", action="store_true")
    args = parser.parse_args()

    if args.merge:
        mesclar_e_gerar()
    else:
        processar_worker(args.worker, args.workers_total)


if __name__ == "__main__":
    main()
