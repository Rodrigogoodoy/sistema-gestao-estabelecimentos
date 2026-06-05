"""
Coleta estabelecimentos SUS de Caxias do Sul/RS e gera CAXIAS_SUS_SIM.xlsx.
"""
import datetime
import pandas as pd
from pathlib import Path
from playwright.sync_api import sync_playwright
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from bot_municipio import processar_municipio
from fix_cnpj_municipios import limpar as limpar_cnpj

URL      = "https://cnes.datasus.gov.br/pages/estabelecimentos/consulta.jsp"
OUTPUT   = Path("resultados_caxias.xlsx")
SAIDA    = "CAXIAS_SUS_SIM.xlsx"
HEADLESS = True

MUNICIPIO = {"cod_ibge": "4305108", "municipio": "Caxias do Sul", "uf": "RS"}

COLUNAS_MAP = {
    "CNPJ":                          "CNPJ",
    "CNPJ Mantenedora":              "CNPJ Mantenedora",
    "Município":                     "Municipio",
    "UF":                            "UF",
    "CNES":                          "Codigo CNES",
    "Nome Fantasia":                 "Nome Fantasia",
    "Nome":                          "Razao Social",
    "Tipo de Estabelecimento":       "Tipo de Estabelecimento",
    "Classificação Estabelecimento": "Classificacao",
    "Gestão":                        "Gestao",
    "Natureza Jurídica(Grupo)":      "Natureza Juridica",
    "CEP":                           "CEP",
    "Logradouro":                    "Logradouro",
    "Número":                        "Numero",
    "Bairro":                        "Bairro",
    "Complemento":                   "Complemento",
    "Telefone":                      "Telefone",
    "E-mail":                        "Email",
    "Responsável Técnico":           "Responsavel Tecnico",
    "Data Desativação":              "Data Desativacao",
    "Motivo Desativação":            "Motivo Desativacao",
}

LARGURAS = {
    "CNPJ": 20, "CNPJ Mantenedora": 22, "Municipio": 22, "UF": 8,
    "Codigo CNES": 14, "Nome Fantasia": 35, "Razao Social": 38,
    "Tipo de Estabelecimento": 30, "Classificacao": 30, "Gestao": 16,
    "Natureza Juridica": 30, "CEP": 14, "Logradouro": 32,
    "Numero": 10, "Bairro": 20, "Complemento": 20,
    "Telefone": 18, "Email": 28, "Responsavel Tecnico": 30,
    "Data Desativacao": 18, "Motivo Desativacao": 40,
}


def _fix_cnpj(row):
    cnpj    = limpar_cnpj(row.get("CNPJ", ""))
    proprio = limpar_cnpj(row.get("CNPJ Próprio", ""))
    mant    = limpar_cnpj(row.get("CNPJ Mantenedora", ""))
    # CNPJ = apenas próprio; não usa mantenedora no lugar
    if cnpj and cnpj == mant and not proprio:
        cnpj = ""
    elif not cnpj and proprio:
        cnpj = proprio
    row["CNPJ"] = cnpj
    return row


def gerar_excel(df: pd.DataFrame):
    cols = {k: v for k, v in COLUNAS_MAP.items() if k in df.columns}
    df_exp = df[list(cols.keys())].rename(columns=cols)

    for col in ("Data Desativacao", "Motivo Desativacao"):
        if col not in df_exp.columns:
            df_exp[col] = ""

    df_exp = df_exp.replace({"null - null": "", "null": "", "---": ""}).fillna("")

    qtd    = len(df_exp)
    ativos = int((df_exp["Data Desativacao"] == "").sum())
    inativos = qtd - ativos
    agora  = datetime.datetime.now().strftime("%d/%m/%Y")

    writer = pd.ExcelWriter(SAIDA, engine="openpyxl")
    df_exp.to_excel(writer, sheet_name="Estabelecimentos", index=False, startrow=4)
    ws = writer.sheets["Estabelecimentos"]

    def fill(cor): return PatternFill("solid", fgColor=cor)
    def borda():
        lado = Side(style="thin", color="D1D5DB")
        return Border(left=lado, right=lado, top=lado, bottom=lado)

    ncols = df_exp.shape[1]
    ult   = get_column_letter(ncols)

    ws.merge_cells(f"A1:{ult}1")
    ws["A1"] = "ESTABELECIMENTOS DE SAUDE — CAXIAS DO SUL/RS — ATENDE SUS: SIM"
    ws["A1"].fill = fill("1E3A8A")
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells(f"A2:{ult}2")
    ws["A2"] = (
        f"Data de geracao: {agora}"
        f"   |   Estabelecimentos: {qtd:,}"
        f"   |   Filtro: Atende SUS = Sim"
        f"   |   Fonte: cnes.datasus.gov.br"
    )
    ws["A2"].fill = fill("2563EB")
    ws["A2"].font = Font(bold=False, color="FFFFFF", size=10, name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    ws.merge_cells(f"A3:{ult}3")
    ws["A3"] = f"Ativos: {ativos:,}   |   Com data de desativação: {inativos:,}"
    ws["A3"].fill = fill("DBEAFE")
    ws["A3"].font = Font(bold=False, color="1E3A8A", size=10, name="Calibri")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 8

    for col in range(1, ncols + 1):
        c = ws.cell(row=5, column=col)
        c.fill = fill("1E3A8A")
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda()
    ws.row_dimensions[5].height = 30

    for row in range(6, qtd + 6):
        bg = "F0F9FF" if row % 2 == 0 else "FFFFFF"
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = fill(bg)
            c.font = Font(bold=False, color="1E293B", size=10, name="Calibri")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = borda()
        ws.row_dimensions[row].height = 18

    for i, col_name in enumerate(df_exp.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = LARGURAS.get(col_name, 18)

    ws.freeze_panes = "A6"
    writer.close()
    print(f"\nGerado: {SAIDA} ({qtd} registros | {ativos} ativos | {inativos} com desativação)")


def main():
    print(f"\nColetando: {MUNICIPIO['municipio']}/{MUNICIPIO['uf']}")
    print("=" * 60)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        page    = browser.new_page()

        resultados = processar_municipio(
            page=page,
            uf=MUNICIPIO["uf"],
            municipio=MUNICIPIO["municipio"],
            cod_ibge=MUNICIPIO["cod_ibge"],
            url_base=URL,
            delay_entre_fichas=0.3,
        )
        browser.close()

    print(f"\nTotal coletado: {len(resultados)} estabelecimentos")

    if not resultados:
        print("Nenhum resultado.")
        return

    df = pd.DataFrame(resultados)
    df = df.apply(_fix_cnpj, axis=1)
    df.to_excel(OUTPUT, index=False)
    print(f"Dados brutos salvos em: {OUTPUT}")

    gerar_excel(df)


if __name__ == "__main__":
    main()
