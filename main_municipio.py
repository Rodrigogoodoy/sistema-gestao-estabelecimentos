"""
Fase 2 — Busca por Município no CNES DataSUS.
Coleta todos os estabelecimentos (Atende SUS: Sim) de cada município.
Gera Excel formatado com visual profissional.
"""
import pandas as pd
from pathlib import Path
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from bot_municipio import processar_municipio

# ── Configurações ──────────────────────────────────────────────────────────────
URL     = "https://cnes.datasus.gov.br/pages/estabelecimentos/consulta.jsp"
OUTPUT  = Path("resultados_fase2.xlsx")
HEADLESS = True
DELAY_ENTRE_FICHAS = 0.3

# ── Municípios ─────────────────────────────────────────────────────────────────
MUNICIPIOS = [
    {"posicao": 23,   "cod_ibge": "3549904", "municipio": "São José dos Campos", "uf": "SP"},
    {"posicao": 192,  "cod_ibge": "4302105", "municipio": "Bento Gonçalves",     "uf": "RS"},
    {"posicao": 320,  "cod_ibge": "4307906", "municipio": "Farroupilha",          "uf": "RS"},
    {"posicao": 403,  "cod_ibge": "4304804", "municipio": "Carlos Barbosa",       "uf": "RS"},
    {"posicao": 434,  "cod_ibge": "4308607", "municipio": "Garibaldi",            "uf": "RS"},
    {"posicao": 461,  "cod_ibge": "4308201", "municipio": "Flores da Cunha",      "uf": "RS"},
    {"posicao": 900,  "cod_ibge": "4319000", "municipio": "São Marcos",           "uf": "RS"},
    {"posicao": 946,  "cod_ibge": "3535309", "municipio": "Palmital",             "uf": "SP"},
    {"posicao": 1092, "cod_ibge": "2706703", "municipio": "Penedo",               "uf": "AL"},
]

# Ordem preferida das colunas (do popup + identificação)
COLUNAS_ORDEM = [
    "UF", "Município", "Cod IBGE",
    "Nome", "CNES", "CNPJ", "Nome Empresarial",
    "Natureza Jurídica(Grupo)",
    "Logradouro", "Número", "Complemento", "Bairro",
    "CEP", "Telefone",
    "Dependência", "Regional de Saúde",
    "Tipo de Estabelecimento", "Subtipo de Estabelecimento", "Gestão",
    "Data Desativação", "Motivo Desativação",
]

# ── Estilos ────────────────────────────────────────────────────────────────────
COR_HEADER    = "1E40AF"
COR_LINHA_PAR = "EFF6FF"
COR_TOTAL     = "DBEAFE"
COR_BORDA     = "CBD5E1"

LARGURAS_FIXAS = {
    "UF": 6, "Município": 20, "Cod IBGE": 11, "CNES": 10,
    "CNPJ": 20, "CEP": 11, "Telefone": 16,
    "Regional de Saúde": 10, "Número": 8, "Complemento": 14,
    "Dependência": 14, "Gestão": 12,
}


def _borda():
    s = Side(style="thin", color=COR_BORDA)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(cor):
    return PatternFill("solid", fgColor=cor)


def _header_cell(cell, texto):
    cell.value     = texto
    cell.fill      = _fill(COR_HEADER)
    cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _borda()


def _data_cell(cell, valor=None, par=False, bold=False, align="left"):
    if valor is not None:
        cell.value = valor
    if par:
        cell.fill = _fill(COR_LINHA_PAR)
    cell.font      = Font(size=10, name="Calibri", bold=bold)
    cell.alignment = Alignment(vertical="center", horizontal=align)
    cell.border    = _borda()


def _ajustar_larguras(ws):
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        header = str(ws.cell(1, col_idx).value or "")
        if header in LARGURAS_FIXAS:
            ws.column_dimensions[get_column_letter(col_idx)].width = LARGURAS_FIXAS[header]
        else:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 48)


def _formatar_aba_dados(ws, n_linhas):
    ws.row_dimensions[1].height = 32
    for cell in ws[1]:
        _header_cell(cell, cell.value)

    for row_idx in range(2, n_linhas + 2):
        ws.row_dimensions[row_idx].height = 16
        par = (row_idx % 2 == 0)
        for cell in ws[row_idx]:
            _data_cell(cell, par=par)

    _ajustar_larguras(ws)
    ws.freeze_panes            = "A2"
    ws.auto_filter.ref         = ws.dimensions
    ws.sheet_view.showGridLines = False


def _criar_aba_resumo(wb, df):
    ws = wb.create_sheet("Resumo")

    headers = ["Município", "UF", "Cod IBGE", "Estabelecimentos"]
    ws.row_dimensions[1].height = 28
    for col_idx, h in enumerate(headers, start=1):
        _header_cell(ws.cell(1, col_idx), h)

    resumo = (
        df.groupby(["Município", "UF", "Cod IBGE"])
        .size()
        .reset_index(name="Estabelecimentos")
        .sort_values("Estabelecimentos", ascending=False)
        .reset_index(drop=True)
    )

    for row_idx, row in resumo.iterrows():
        r = int(row_idx) + 2
        par = (r % 2 == 0)
        ws.row_dimensions[r].height = 16
        _data_cell(ws.cell(r, 1), row["Município"],            par=par)
        _data_cell(ws.cell(r, 2), row["UF"],                   par=par, align="center")
        _data_cell(ws.cell(r, 3), str(row["Cod IBGE"]),        par=par, align="center")
        _data_cell(ws.cell(r, 4), int(row["Estabelecimentos"]),par=par, bold=True, align="right")

    # Linha de total
    total_row = len(resumo) + 2
    ws.row_dimensions[total_row].height = 18
    for col_idx in range(1, 5):
        cell = ws.cell(total_row, col_idx)
        cell.fill   = _fill(COR_TOTAL)
        cell.border = _borda()
        cell.font   = Font(bold=True, size=10, name="Calibri", color=COR_HEADER)
        cell.alignment = Alignment(vertical="center",
                                   horizontal="right" if col_idx == 4 else "left")
    ws.cell(total_row, 1).value = "TOTAL"
    ws.cell(total_row, 4).value = int(resumo["Estabelecimentos"].sum())
    ws.cell(total_row, 4).font  = Font(bold=True, size=12, name="Calibri", color=COR_HEADER)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.freeze_panes            = "A2"
    ws.auto_filter.ref         = f"A1:D{len(resumo) + 1}"
    ws.sheet_view.showGridLines = False


def _salvar_excel(resultados: list, caminho: Path):
    if not resultados:
        print("Nenhum resultado para salvar.")
        return

    df = pd.DataFrame(resultados)

    # Ordena colunas
    presentes = [c for c in COLUNAS_ORDEM if c in df.columns]
    extras    = [c for c in df.columns if c not in presentes]
    df = df[presentes + extras]

    # Salva base com pandas
    df.to_excel(caminho, index=False, sheet_name="Estabelecimentos", engine="openpyxl")

    # Abre e formata
    wb = load_workbook(caminho)
    _formatar_aba_dados(wb["Estabelecimentos"], len(df))
    _criar_aba_resumo(wb, df)
    wb.save(caminho)

    print(f"\nSalvo: {caminho} ({len(df)} registros)")


def main():
    todos = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        page    = browser.new_page()

        for m in MUNICIPIOS:
            print(
                f"\n{'='*60}\n"
                f"Município: {m['municipio']}/{m['uf']}  |  IBGE: {m['cod_ibge']}\n"
                f"{'='*60}"
            )
            try:
                resultados = processar_municipio(
                    page=page,
                    uf=m["uf"],
                    municipio=m["municipio"],
                    cod_ibge=m["cod_ibge"],
                    url_base=URL,
                    delay_entre_fichas=DELAY_ENTRE_FICHAS,
                )
                todos.extend(resultados)
                print(f"  Concluído: {len(resultados)} estabelecimentos")
            except Exception as e:
                print(f"  ERRO GERAL em {m['municipio']}/{m['uf']}: {e}")

        browser.close()

    _salvar_excel(todos, OUTPUT)

    print("\n" + "=" * 60)
    print(f"TOTAL GERAL: {len(todos)} estabelecimentos")
    por_municipio = {}
    for r in todos:
        chave = f"{r.get('Município','?')}/{r.get('UF','?')}"
        por_municipio[chave] = por_municipio.get(chave, 0) + 1
    for mun, qtd in sorted(por_municipio.items(), key=lambda x: -x[1]):
        print(f"  {mun}: {qtd}")
    print("=" * 60)


if __name__ == "__main__":
    main()
