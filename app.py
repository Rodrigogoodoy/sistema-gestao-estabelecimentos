"""
Sistema CNES DataSUS — Servidor Web
Interface para pesquisar estabelecimentos SUS por município.
"""
import threading
import uuid
import datetime
import os
import io
import json
from functools import wraps
from pathlib import Path

import requests as req
import pandas as pd
from flask import Flask, jsonify, request, send_file, session, redirect
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

from bot_municipio import processar_municipio

app = Flask(__name__, static_folder=".", static_url_path="")
app.secret_key = "cnes-gestao-saude-rodrigo-2025-chave"

# ── Usuários ─────────────────────────────────────────────────────────────────
USUARIOS_FILE  = Path("usuarios.json")
_usuarios_lock = threading.Lock()


def _carregar_usuarios():
    with _usuarios_lock:
        if not USUARIOS_FILE.exists():
            dados = {"admin": {"senha_hash": generate_password_hash("LUCIANO1234"),
                                "role": "admin", "nome": "Administrador"}}
            with open(USUARIOS_FILE, "w", encoding="utf-8") as f:
                json.dump(dados, f, ensure_ascii=False, indent=2)
            print("  usuarios.json criado (admin / LUCIANO1234)")
            return dados
        with open(USUARIOS_FILE, encoding="utf-8") as f:
            return json.load(f)


def _salvar_usuarios(dados):
    with _usuarios_lock:
        with open(USUARIOS_FILE, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=2)


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "usuario" not in session:
            if request.path.startswith("/api/"):
                return jsonify({"error": "nao_autenticado"}), 401
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "usuario" not in session:
            return jsonify({"error": "nao_autenticado"}), 401
        if session.get("role") != "admin":
            return jsonify({"error": "acesso_negado"}), 403
        return f(*args, **kwargs)
    return decorated


_municipios_cache = None
_municipios_lock  = threading.Lock()
_jobs             = {}
_jobs_lock        = threading.Lock()

COLUNAS_MAP = {
    "CNPJ":                          "CNPJ",
    "CNPJ Mantenedora":              "CNPJ Mantenedora",
    "Município":                     "Municipio",
    "UF":                            "UF",
    "CNES":                          "Codigo CNES",
    "Nome Fantasia":                 "Nome Fantasia",
    "Nome":                          "Razao Social",
    "Tipo de Estabelecimento":       "Tipo de Estabelecimento",
    "Classificação Estabelecimento": "Classificacao",
    "Gestão":                        "Gestao",
    "Natureza Jurídica(Grupo)":      "Natureza Juridica",
    "CEP":                           "CEP",
    "Logradouro":                    "Logradouro",
    "Número":                        "Numero",
    "Bairro":                        "Bairro",
    "Complemento":                   "Complemento",
    "Telefone":                      "Telefone",
    "E-mail":                        "Email",
    "Responsável Técnico":           "Responsavel Tecnico",
    "Data Desativação":              "Data Desativacao",
    "Motivo Desativação":            "Motivo Desativacao",
}

LARGURAS = {
    "CNPJ": 20, "CNPJ Mantenedora": 22, "Municipio": 22, "UF": 8,
    "Codigo CNES": 14, "Nome Fantasia": 35, "Razao Social": 38,
    "Tipo de Estabelecimento": 30, "Classificacao": 30, "Gestao": 16,
    "Natureza Juridica": 30, "CEP": 14, "Logradouro": 32,
    "Numero": 10, "Bairro": 20, "Complemento": 20,
    "Telefone": 18, "Email": 28, "Responsavel Tecnico": 30,
    "Data Desativacao": 18, "Motivo Desativacao": 40,
}


def _carregar_municipios():
    global _municipios_cache
    with _municipios_lock:
        if _municipios_cache is not None:
            return _municipios_cache
        # Tenta carregar do arquivo local primeiro
        local = Path("municipios.json")
        if local.exists():
            print("Carregando municípios do arquivo local...")
            import json
            with open(local, encoding="utf-8") as f:
                _municipios_cache = json.load(f)
            total = sum(len(v) for v in _municipios_cache.values())
            print(f"  {total} municípios carregados ({len(_municipios_cache)} estados).")
            return _municipios_cache
        # Fallback: API do IBGE
        print("Carregando municípios do IBGE...")
        try:
            resp = req.get(
                "https://servicodados.ibge.gov.br/api/v1/localidades/municipios?orderBy=nome",
                timeout=30,
            )
            dados = resp.json()
            result = {}
            for m in dados:
                try:
                    uf = m["microrregiao"]["mesorregiao"]["UF"]["sigla"]
                    if uf not in result:
                        result[uf] = []
                    result[uf].append({"nome": m["nome"], "cod_ibge": str(m["id"])})
                except Exception:
                    pass
            _municipios_cache = result
            total = sum(len(v) for v in result.values())
            print(f"  {total} municípios carregados ({len(result)} estados).")
        except Exception as e:
            print(f"  Erro ao carregar municípios: {e}")
            _municipios_cache = {}
        return _municipios_cache


def _limpar_cnpj(v):
    v = str(v or "").strip()
    return "" if v in ("---", "-", "nan", "None", "NaN") else v


def _fix_cnpj(row):
    cnpj    = _limpar_cnpj(row.get("CNPJ", ""))
    proprio = _limpar_cnpj(row.get("CNPJ Próprio", ""))
    mant    = _limpar_cnpj(row.get("CNPJ Mantenedora", ""))
    if cnpj and cnpj == mant and not proprio:
        cnpj = ""
    elif not cnpj and proprio:
        cnpj = proprio
    row["CNPJ"] = cnpj
    return row


def _gerar_excel(df, municipio, uf, output_path):
    cols   = {k: v for k, v in COLUNAS_MAP.items() if k in df.columns}
    df_exp = df[list(cols.keys())].rename(columns=cols)
    for col in ("Data Desativacao", "Motivo Desativacao"):
        if col not in df_exp.columns:
            df_exp[col] = ""
    df_exp = df_exp.replace({"null - null": "", "null": "", "---": ""}).fillna("")

    qtd      = len(df_exp)
    ativos   = int((df_exp["Data Desativacao"] == "").sum())
    inativos = qtd - ativos

    writer = pd.ExcelWriter(output_path, engine="openpyxl")
    df_exp.to_excel(writer, sheet_name="Estabelecimentos", index=False, startrow=0)
    ws = writer.sheets["Estabelecimentos"]

    def fill(cor): return PatternFill("solid", fgColor=cor)
    def borda():
        lado = Side(style="thin", color="D1D5DB")
        return Border(left=lado, right=lado, top=lado, bottom=lado)

    ncols = df_exp.shape[1]

    # Cabeçalho das colunas na linha 1
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill("1E3A8A")
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda()
    ws.row_dimensions[1].height = 30

    # Dados a partir da linha 2
    for row in range(2, qtd + 2):
        bg = "F0F9FF" if row % 2 == 0 else "FFFFFF"
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = fill(bg)
            c.font = Font(bold=False, color="1E293B", size=10, name="Calibri")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = borda()
        ws.row_dimensions[row].height = 18

    for i, col_name in enumerate(df_exp.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = LARGURAS.get(col_name, 18)

    ws.freeze_panes = "A2"
    writer.close()
    return qtd, ativos, inativos


def _run_scrape(job_id, uf, municipio, cod_ibge, auto_open=True):
    job = _jobs[job_id]
    job["log"] = []

    def log(msg):
        job["log"].append(msg)
        print(f"  [{job_id}] {msg}")

    try:
        log(f"Iniciando coleta: {municipio}/{uf}")
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page    = browser.new_page()
            resultados = processar_municipio(
                page=page,
                uf=uf,
                municipio=municipio,
                cod_ibge=cod_ibge,
                url_base="https://cnes.datasus.gov.br/pages/estabelecimentos/consulta.jsp",
                delay_entre_fichas=0.3,
            )
            browser.close()

        if not resultados:
            job["status"] = "error"
            job["error"]  = "Nenhum estabelecimento encontrado para este município."
            return

        log(f"{len(resultados)} estabelecimentos coletados. Gerando Excel...")

        df = pd.DataFrame(resultados)
        df = df.apply(_fix_cnpj, axis=1)

        safe = municipio.replace(" ", "_").replace("/", "_")
        pasta = Path("resultados")
        pasta.mkdir(exist_ok=True)
        output_path = pasta / f"resultado_{safe}_{uf}_{job_id}.xlsx"
        qtd, ativos, inativos = _gerar_excel(df, municipio, uf, str(output_path))

        job["status"]   = "done"
        job["file"]     = str(output_path)
        job["filename"] = f"{safe}_{uf}_SUS_SIM.xlsx"
        job["qtd"]      = qtd
        job["ativos"]   = ativos
        job["inativos"] = inativos
        log(f"Concluído: {qtd} registros | {ativos} ativos | {inativos} desativados")

        if auto_open and os.name == "nt":
            os.startfile(str(output_path.resolve()))

    except Exception as e:
        job["status"] = "error"
        job["error"]  = str(e)
        log(f"Erro: {e}")


# ── Rotas ────────────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def index():
    return app.send_static_file("index.html")


@app.route("/login")
def login_page():
    if "usuario" in session:
        return redirect("/")
    return app.send_static_file("login.html")


@app.route("/admin")
@login_required
def admin_page():
    if session.get("role") != "admin":
        return redirect("/")
    return app.send_static_file("admin.html")


@app.route("/api/login", methods=["POST"])
def api_login():
    body    = request.get_json() or {}
    usuario = body.get("usuario", "").strip().lower()
    senha   = body.get("senha", "")
    dados   = _carregar_usuarios()
    user    = dados.get(usuario)
    if not user or not check_password_hash(user["senha_hash"], senha):
        return jsonify({"error": "Usuário ou senha incorretos"}), 401
    session["usuario"] = usuario
    session["role"]    = user.get("role", "user")
    session["nome"]    = user.get("nome", usuario)
    return jsonify({"ok": True, "role": session["role"], "nome": session["nome"]})


@app.route("/api/logout")
def api_logout():
    session.clear()
    return redirect("/login")


@app.route("/api/me")
@login_required
def api_me():
    return jsonify({"usuario": session["usuario"], "role": session["role"], "nome": session["nome"]})


@app.route("/api/admin/usuarios")
@admin_required
def admin_listar():
    dados = _carregar_usuarios()
    return jsonify([{"usuario": k, "nome": v.get("nome", k), "role": v.get("role", "user")}
                    for k, v in dados.items()])


@app.route("/api/admin/criar", methods=["POST"])
@admin_required
def admin_criar():
    body    = request.get_json() or {}
    usuario = body.get("usuario", "").strip().lower()
    nome    = body.get("nome", "").strip()
    senha   = body.get("senha", "").strip()
    role    = body.get("role", "user")
    if not usuario or not nome or not senha:
        return jsonify({"error": "Usuário, nome e senha são obrigatórios"}), 400
    if role not in ("admin", "user"):
        return jsonify({"error": "Role inválida"}), 400
    dados = _carregar_usuarios()
    if usuario in dados:
        return jsonify({"error": "Usuário já existe"}), 409
    dados[usuario] = {"senha_hash": generate_password_hash(senha), "role": role, "nome": nome}
    _salvar_usuarios(dados)
    return jsonify({"ok": True})


@app.route("/api/admin/deletar/<username>", methods=["DELETE"])
@admin_required
def admin_deletar(username):
    if username == session["usuario"]:
        return jsonify({"error": "Não é possível deletar seu próprio usuário"}), 400
    dados = _carregar_usuarios()
    if username not in dados:
        return jsonify({"error": "Usuário não encontrado"}), 404
    del dados[username]
    _salvar_usuarios(dados)
    return jsonify({"ok": True})


@app.route("/api/admin/senha/<username>", methods=["PUT"])
@admin_required
def admin_senha(username):
    body  = request.get_json() or {}
    senha = body.get("senha", "").strip()
    if len(senha) < 4:
        return jsonify({"error": "Senha deve ter ao menos 4 caracteres"}), 400
    dados = _carregar_usuarios()
    if username not in dados:
        return jsonify({"error": "Usuário não encontrado"}), 404
    dados[username]["senha_hash"] = generate_password_hash(senha)
    _salvar_usuarios(dados)
    return jsonify({"ok": True})


@app.route("/api/municipios/<uf>")
@login_required
def get_municipios_uf(uf):
    data = _carregar_municipios()
    return jsonify(data.get(uf.upper(), []))


@app.route("/api/pesquisar", methods=["POST"])
@login_required
def pesquisar():
    body     = request.get_json() or {}
    uf       = body.get("uf", "").strip().upper()
    municipio = body.get("municipio", "").strip()
    cod_ibge  = body.get("cod_ibge", "").strip()

    if not uf or not municipio or not cod_ibge:
        return jsonify({"error": "UF, município e código IBGE são obrigatórios."}), 400

    auto_open = not body.get("lote", False)

    job_id = str(uuid.uuid4())[:8]
    with _jobs_lock:
        _jobs[job_id] = {
            "status": "running", "file": None, "filename": None,
            "error": None, "qtd": 0, "ativos": 0, "inativos": 0, "log": [],
        }

    threading.Thread(target=_run_scrape, args=(job_id, uf, municipio, cod_ibge, auto_open), daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/api/status/<job_id>")
@login_required
def get_status(job_id):
    with _jobs_lock:
        job = _jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job não encontrado"}), 404
    return jsonify(job)


@app.route("/api/download/<job_id>")
@login_required
def download(job_id):
    with _jobs_lock:
        job = _jobs.get(job_id)
    if not job or not job.get("file"):
        return jsonify({"error": "Arquivo não disponível"}), 404
    return send_file(
        job["file"],
        as_attachment=True,
        download_name=job.get("filename", "resultado.xlsx"),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/download/lote", methods=["POST"])
@login_required
def download_lote():
    body    = request.get_json() or {}
    job_ids = body.get("job_ids", [])
    if not job_ids:
        return jsonify({"error": "Nenhum job fornecido"}), 400

    with _jobs_lock:
        jobs = [_jobs[jid] for jid in job_ids if jid in _jobs and _jobs[jid].get("status") == "done" and _jobs[jid].get("file")]

    if not jobs:
        return jsonify({"error": "Nenhum resultado disponível"}), 404

    # Lê cada Excel e combina — os arquivos têm 4 linhas de cabeçalho antes dos dados
    dfs = []
    for job in jobs:
        try:
            df_job = pd.read_excel(job["file"], sheet_name="Estabelecimentos", header=0)
            dfs.append(df_job)
        except Exception as e:
            print(f"  [lote] Erro ao ler {job['file']}: {e}")

    if not dfs:
        return jsonify({"error": "Erro ao ler arquivos de resultado"}), 500

    df_combined = pd.concat(dfs, ignore_index=True).fillna("")

    buf = io.BytesIO()
    df_combined.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name="lote_sus_sim.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/abrir/lote", methods=["POST"])
@login_required
def abrir_lote():
    body    = request.get_json() or {}
    job_ids = body.get("job_ids", [])
    if not job_ids:
        return jsonify({"error": "Nenhum job fornecido"}), 400

    with _jobs_lock:
        jobs = [_jobs[jid] for jid in job_ids if jid in _jobs and _jobs[jid].get("status") == "done" and _jobs[jid].get("file")]

    if not jobs:
        return jsonify({"error": "Nenhum resultado disponível"}), 404

    dfs = []
    for job in jobs:
        try:
            df_job = pd.read_excel(job["file"], sheet_name="Estabelecimentos", header=0)
            dfs.append(df_job)
        except Exception as e:
            print(f"  [lote] Erro ao ler {job['file']}: {e}")

    if not dfs:
        return jsonify({"error": "Erro ao ler arquivos"}), 500

    df_combined = pd.concat(dfs, ignore_index=True).fillna("")

    pasta = Path("resultados")
    pasta.mkdir(exist_ok=True)
    output_path = pasta / f"lote_combinado_{uuid.uuid4().hex[:6]}.xlsx"

    writer = pd.ExcelWriter(str(output_path), engine="openpyxl")
    df_combined.to_excel(writer, sheet_name="Estabelecimentos", index=False, startrow=0)
    ws = writer.sheets["Estabelecimentos"]

    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def fill(cor): return PatternFill("solid", fgColor=cor)
    lado = Side(style="thin", color="D1D5DB")
    def borda(): return Border(left=lado, right=lado, top=lado, bottom=lado)

    ncols = df_combined.shape[1]
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill("1E3A8A")
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda()
    ws.row_dimensions[1].height = 30

    for row in range(2, len(df_combined) + 2):
        bg = "F0F9FF" if row % 2 == 0 else "FFFFFF"
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = fill(bg)
            c.font = Font(bold=False, color="1E293B", size=10, name="Calibri")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = borda()
        ws.row_dimensions[row].height = 18

    for i, col_name in enumerate(df_combined.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = LARGURAS.get(col_name, 18)

    ws.freeze_panes = "A2"
    writer.close()

    if os.name == "nt":
        os.startfile(str(output_path.resolve()))

    return jsonify({"ok": True, "total": len(df_combined)})


def _iniciar_cloudflare_tunnel():
    """Inicia cloudflared tunnel e retorna a URL pública."""
    import subprocess, re, time
    cf = Path(__file__).parent / "cloudflared.exe"
    if not cf.exists():
        return None
    try:
        proc = subprocess.Popen(
            [str(cf), "tunnel", "--url", "http://localhost:5000"],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
        )
        # Aguarda até 20s pela URL
        for _ in range(40):
            line = proc.stdout.readline()
            m = re.search(r"https://[a-zA-Z0-9\-]+\.trycloudflare\.com", line)
            if m:
                return m.group(0)
            time.sleep(0.5)
    except Exception as e:
        print(f"  Cloudflare Tunnel erro: {e}")
    return None


if __name__ == "__main__":
    _carregar_municipios()

    print()
    print("=" * 60)
    print("  Iniciando túnel público (Cloudflare)...")
    public_url = _iniciar_cloudflare_tunnel()

    # Salva URL em arquivo para consulta posterior
    with open("tunnel_url.txt", "w", encoding="utf-8") as f:
        f.write(public_url if public_url else "nao_disponivel")

    print("=" * 60)
    print("  SISTEMA CNES DataSUS — SERVIDOR ATIVO")
    print("=" * 60)
    print(f"  Acesso local:   http://localhost:5000")
    if public_url:
        print(f"  Acesso publico: {public_url}")
        print(f"")
        print(f"  Compartilhe esse link com qualquer pessoa!")
    else:
        print("  Acesso publico: nao disponivel (cloudflared.exe nao encontrado)")
    print("=" * 60)
    print()

    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
