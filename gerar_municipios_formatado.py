"""
Gera cidades_pequenas_por_municipio.xlsx com uma aba por municipio,
usando a mesma formatacao de tabela azul do arquivo original.
"""

import pandas as pd
import urllib.request
import json
import gzip
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

AZUL_HEADER = "1F497D"
STYLE_TABELA = "TableStyleMedium2"
LIMITE_POP   = 50_000
ARQUIVO_SAIDA = "cidades_pequenas_por_municipio.xlsx"

# ------------------------------------------------------------------ #
# Populacao IBGE
# ------------------------------------------------------------------ #
def buscar_populacao():
    print("Buscando populacao IBGE...")
    url = (
        "https://servicodados.ibge.gov.br/api/v3/agregados/6579"
        "/periodos/2021/variaveis/9324?localidades=N6[all]"
    )
    req = urllib.request.Request(url, headers={"Accept-Encoding": "identity"})
    with urllib.request.urlopen(req, timeout=60) as r:
        raw = r.read()
        if raw[:2] == b"\x1f\x8b":
            raw = gzip.decompress(raw)
        dados = json.loads(raw.decode("utf-8"))
    pop = {}
    for item in dados[0]["resultados"][0]["series"]:
        cod = item["localidade"]["id"][:6]
        val = item["serie"].get("2021")
        if val and val != "-":
            try:
                pop[cod] = int(val)
            except ValueError:
                pass
    print(f"  {len(pop)} municipios carregados.")
    return pop

# ------------------------------------------------------------------ #
# Formatar aba
# ------------------------------------------------------------------ #
def formatar_aba(ws, n_linhas, n_colunas):
    fill_azul  = PatternFill("solid", fgColor=AZUL_HEADER)
    fonte_hdr  = Font(bold=True, color="FFFFFF")
    alinhamento = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Header
    for cell in ws[1]:
        cell.fill  = fill_azul
        cell.font  = fonte_hdr
        cell.alignment = alinhamento

    # Altura do header
    ws.row_dimensions[1].height = 20

    # Tabela Excel (range de dados)
    ultima_col = get_column_letter(n_colunas)
    ref = f"A1:{ultima_col}{n_linhas + 1}"
    tabela = Table(displayName=f"T_{ws.title.replace('-','_').replace(' ','_')[:28]}", ref=ref)
    tabela.tableStyleInfo = TableStyleInfo(
        name=STYLE_TABELA,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabela)

    # Auto-largura das colunas (limite 50)
    for col_idx in range(1, n_colunas + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(cell.value)) for cell in ws[col_letter] if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

# ------------------------------------------------------------------ #
# Main
# ------------------------------------------------------------------ #
def main():
    pop = buscar_populacao()

    print("Carregando resultados_SIM_final.xlsx...")
    df = pd.read_excel("resultados_SIM_final.xlsx")

    col_mun = "Município" if "Município" in df.columns else "Municipio"
    df["_ibge6"] = df[col_mun].astype(str).str.extract(r"^(\d{6})")
    df["_cidade"] = df[col_mun].astype(str).str.extract(r"^\d+ - (.+)$")

    codigos_peq = {cod for cod, p in pop.items() if p <= LIMITE_POP}
    df_peq = (
        df[df["_ibge6"].isin(codigos_peq)]
        .copy()
        .sort_values(["UF", "_cidade", "CNPJ Pesquisado"])
    )
    colunas_saida = [c for c in df_peq.columns if not c.startswith("_")]
    df_peq = df_peq[colunas_saida]

    print(f"  {len(df_peq)} registros em cidades pequenas.")

    # Lista de municipios ordenados
    municipios = (
        df_peq[["UF", col_mun]]
        .drop_duplicates()
        .sort_values(["UF", col_mun])
        .values.tolist()
    )

    def sanitizar(nome):
        nome = re.sub(r"[\\/*?:\[\]]", "", str(nome).strip())
        return nome[:31]

    print(f"Gerando {len(municipios)} abas em {ARQUIVO_SAIDA}...")

    # Escrever com pandas primeiro (rapido)
    with pd.ExcelWriter(ARQUIVO_SAIDA, engine="openpyxl") as writer:
        # Aba indice
        resumo = (
            df_peq.groupby(["UF", col_mun])
            .size()
            .reset_index(name="Qtd de CNPJs")
            .rename(columns={col_mun: "Município"})
        )
        resumo.to_excel(writer, sheet_name="ÍNDICE", index=False)

        for i, (uf, mun) in enumerate(municipios, 1):
            aba = sanitizar(f"{uf}-{mun}")
            df_mun = df_peq[
                (df_peq["UF"] == uf) & (df_peq[col_mun] == mun)
            ]
            df_mun.to_excel(writer, sheet_name=aba, index=False)
            if i % 200 == 0:
                print(f"  {i}/{len(municipios)} abas gravadas...")

    print("Aplicando formatacao...")
    wb = load_workbook(ARQUIVO_SAIDA)

    # Formatar indice
    ws_idx = wb["ÍNDICE"]
    formatar_aba(ws_idx, len(resumo), len(resumo.columns))

    # Formatar cada municipio
    for i, (uf, mun) in enumerate(municipios, 1):
        aba = sanitizar(f"{uf}-{mun}")
        if aba not in wb.sheetnames:
            continue
        ws = wb[aba]
        n_linhas = ws.max_row - 1
        n_colunas = ws.max_column
        formatar_aba(ws, n_linhas, n_colunas)
        if i % 200 == 0:
            print(f"  {i}/{len(municipios)} abas formatadas...")

    wb.save(ARQUIVO_SAIDA)
    print(f"\nConcluido! Arquivo salvo: {ARQUIVO_SAIDA}")
    print(f"Total: {len(municipios)} abas de municipios + 1 aba INDICE")


if __name__ == "__main__":
    main()
