import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

df = pd.read_excel('resultados.xlsx', dtype=str)
sim = df[df['ENCONTRADO'] == 'SIM'].copy().reset_index(drop=True)
total_pesquisados = len(df)
qtd_sim = len(sim)
pct_sim = round(qtd_sim / total_pesquisados * 100, 1)
agora = datetime.datetime.now().strftime("%d/%m/%Y")

# Colunas relevantes e seus nomes amigaveis
colunas_map = {
    'CNPJ Pesquisado':          'CNPJ',
    'CNES':                     'Codigo CNES',
    'Nome Fantasia':            'Nome Fantasia',
    'Nome':                     'Razao Social',
    'Tipo de Estabelecimento':  'Tipo de Estabelecimento',
    'Classificação Estabelecimento': 'Classificacao',
    'Gestão':                   'Gestao',
    'Natureza Jurídica(Grupo)': 'Natureza Juridica',
    'Municipio':                'Municipio',
    'UF':                       'UF',
    'CEP':                      'CEP',
    'Logradouro':               'Logradouro',
    'Número':                   'Numero',
    'Bairro':                   'Bairro',
    'Telefone':                 'Telefone',
    'E-mail':                   'Email',
    'Responsável Técnico':      'Responsavel Tecnico',
}

# Pega apenas colunas que existem no dataframe
colunas_existentes = {k: v for k, v in colunas_map.items() if k in sim.columns}
sim_export = sim[list(colunas_existentes.keys())].rename(columns=colunas_existentes)

writer = pd.ExcelWriter('ESTABELECIMENTOS_CNES.xlsx', engine='openpyxl')
sim_export.to_excel(writer, sheet_name='Estabelecimentos', index=False, startrow=4)
ws = writer.sheets['Estabelecimentos']

def fill(cor):   return PatternFill('solid', fgColor=cor)
def font(bold=False, color='1E293B', size=11): return Font(bold=bold, color=color, size=size, name='Calibri')
def borda():
    lado = Side(style='thin', color='D1D5DB')
    return Border(left=lado, right=lado, top=lado, bottom=lado)

ncols = sim_export.shape[1]
ultima_col = get_column_letter(ncols)

# --- Linha 1: Titulo principal ---
ws.merge_cells('A1:' + ultima_col + '1')
ws['A1'] = 'ESTABELECIMENTOS DE SAUDE ENCONTRADOS NO CNES / DATASUS'
ws['A1'].fill = fill('1E3A8A')
ws['A1'].font = Font(bold=True, color='FFFFFF', size=14, name='Calibri')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 38

# --- Linha 2: Subtitulo com data e stats ---
ws.merge_cells('A2:' + ultima_col + '2')
ws['A2'] = (
    'Data de geracao: ' + agora +
    '   |   CNPJs pesquisados: ' + str(total_pesquisados) +
    '   |   Encontrados: ' + str(qtd_sim) +
    ' (' + str(pct_sim) + '%)' +
    '   |   Fonte: cnes.datasus.gov.br'
)
ws['A2'].fill = fill('2563EB')
ws['A2'].font = Font(bold=False, color='FFFFFF', size=10, name='Calibri')
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 22

# --- Linha 3: Nota ---
ws.merge_cells('A3:' + ultima_col + '3')
ws['A3'] = 'Lista com ' + str(qtd_sim) + ' estabelecimentos confirmados no cadastro nacional de saude (CNES)'
ws['A3'].fill = fill('DBEAFE')
ws['A3'].font = Font(bold=False, color='1E3A8A', size=10, name='Calibri')
ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[3].height = 20

# --- Linha 4: espaco ---
ws.row_dimensions[4].height = 8

# --- Linha 5: Cabecalho da tabela ---
for col in range(1, ncols + 1):
    c = ws.cell(row=5, column=col)
    c.fill  = fill('1E3A8A')
    c.font  = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = borda()
ws.row_dimensions[5].height = 30

# --- Linhas de dados ---
for row in range(6, qtd_sim + 6):
    bg = 'F0F9FF' if row % 2 == 0 else 'FFFFFF'
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill      = fill(bg)
        c.font      = Font(bold=False, color='1E293B', size=10, name='Calibri')
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border    = borda()
    ws.row_dimensions[row].height = 18

# --- Largura das colunas ---
larguras = {
    'CNPJ':                    20,
    'Codigo CNES':             14,
    'Nome Fantasia':           35,
    'Razao Social':            38,
    'Tipo de Estabelecimento': 30,
    'Classificacao':           30,
    'Gestao':                  16,
    'Natureza Juridica':       30,
    'Municipio':               22,
    'UF':                       8,
    'CEP':                     14,
    'Logradouro':              32,
    'Numero':                  10,
    'Bairro':                  20,
    'Telefone':                18,
    'Email':                   28,
    'Responsavel Tecnico':     30,
}
for i, col_name in enumerate(sim_export.columns, 1):
    ws.column_dimensions[get_column_letter(i)].width = larguras.get(col_name, 20)

# --- Congelar painel no cabecalho ---
ws.freeze_panes = 'A6'

writer.close()
print('ESTABELECIMENTOS_CNES.xlsx gerado!')
print(str(qtd_sim) + ' estabelecimentos encontrados (' + str(pct_sim) + '% de ' + str(total_pesquisados) + ' pesquisados)')
