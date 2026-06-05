"""
Gera cnpjs_municipios_pequenos.xlsx:
  - Fonte  : resultados_SIM_final.xlsx (estabelecimentos coletados do CNES)
  - Filtro : apenas CNPJs validos (14 digitos) - exclui CPF / pessoa fisica
  - Filtro : municipios com ate 50.000 habitantes (IBGE 2021)
  - Saida  : 1 aba INDICE  +  1 aba por municipio  (formatacao azul)
"""

import re
import json
import gzip
import urllib.request
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ARQUIVO_ENTRADA = "resultados_SIM_final.xlsx"
ARQUIVO_SAIDA   = "cnpjs_municipios_pequenos.xlsx"
LIMITE_POP      = 50_000

AZUL_HEADER  = "1F497D"
STYLE_TABELA = "TableStyleMedium2"

# Colunas uteis para o relatorio final (na ordem desejada)
COLUNAS_SAIDA = [
    "CNPJ Pesquisado",
    "CNES",
    "Nome Fantasia",
    "Nome Empresarial",
    "Tipo de Estabelecimento",
    "Classificação Estabelecimento",
    "Gestão",
    "Natureza Jurídica(Grupo)",
    "Município",
    "UF",
    "CEP",
    "Logradouro",
    "Número",
    "Bairro",
    "Complemento",
    "Telefone",
    "E-mail",
    "Responsável Técnico",
    "Latitude",
    "Longitude",
]


# ------------------------------------------------------------------ #
# Populacao IBGE
# ------------------------------------------------------------------ #
def buscar_populacao():
    print("Buscando populacao IBGE 2021...")
    url = (
        "https://servicodados.ibge.gov.br/api/v3/agregados/6579"
        "/periodos/2021/variaveis/9324?localidades=N6[all]"
    )
    req = urllib.request.Request(url, headers={"Accept-Encoding": "identity"})
    with urllib.request.urlopen(req, timeout=60) as r:
        raw = r.read()
        if raw[:2] == b"\x1f\x8b":
            raw = gzip.decompress(raw)
        dados = json.loads(raw.decode("utf-8"))
    pop = {}
    for item in dados[0]["resultados"][0]["series"]:
        cod = item["localidade"]["id"][:6]
        val = item["serie"].get("2021")
        if val and val != "-":
            try:
                pop[cod] = int(val)
            except ValueError:
                pass
    print(f"  {len(pop)} municipios com dados populacionais.")
    return pop


# ------------------------------------------------------------------ #
# Helpers
# ------------------------------------------------------------------ #
def _achar_coluna(df, *candidatos):
    """Retorna o primeiro nome de coluna que existir no DataFrame."""
    for nome in candidatos:
        if nome in df.columns:
            return nome
    # Busca parcial (ignora acentos / maiusculas)
    lower = {c.lower().replace("ã", "a").replace("í", "i").replace("ú", "u"): c
             for c in df.columns}
    for nome in candidatos:
        chave = nome.lower().replace("ã", "a").replace("í", "i").replace("ú", "u")
        if chave in lower:
            return lower[chave]
    raise KeyError(f"Nenhuma das colunas encontrada: {candidatos}")


def sanitizar(nome):
    nome = re.sub(r'[\\/*?:\[\]]', "", str(nome).strip())
    return nome[:31]


# ------------------------------------------------------------------ #
# Formatacao Excel
# ------------------------------------------------------------------ #
def formatar_aba(ws, n_linhas, n_colunas):
    fill_azul  = PatternFill("solid", fgColor=AZUL_HEADER)
    fonte_hdr  = Font(bold=True, color="FFFFFF")
    alinhamento = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill      = fill_azul
        cell.font      = fonte_hdr
        cell.alignment = alinhamento
    ws.row_dimensions[1].height = 20

    ultima_col = get_column_letter(n_colunas)
    ref = f"A1:{ultima_col}{n_linhas + 1}"
    nome_tabela = "T_" + re.sub(r"[^A-Za-z0-9_]", "_", ws.title)[:28]
    tabela = Table(displayName=nome_tabela, ref=ref)
    tabela.tableStyleInfo = TableStyleInfo(
        name=STYLE_TABELA,
        showRowStripes=True,
        showColumnStripes=False,
        showFirstColumn=False,
        showLastColumn=False,
    )
    ws.add_table(tabela)

    for col_idx in range(1, n_colunas + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(cell.value)) for cell in ws[col_letter] if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


# ------------------------------------------------------------------ #
# Main
# ------------------------------------------------------------------ #
def main():
    # 1. Populacao
    pop = buscar_populacao()

    # 2. Carregar dados
    print(f"\nCarregando {ARQUIVO_ENTRADA}...")
    df = pd.read_excel(ARQUIVO_ENTRADA, dtype=str)
    print(f"  {len(df)} registros carregados.")

    # 3. Filtrar pessoa fisica: manter so CNPJs validos (14 digitos)
    df["_cnpj_limpo"] = df["CNPJ Pesquisado"].astype(str).str.replace(r"\D", "", regex=True)
    antes = len(df)
    df = df[df["_cnpj_limpo"].str.len() == 14].copy()
    if antes != len(df):
        print(f"  Removidas {antes - len(df)} entradas de pessoa fisica (CPF / invalido).")
    else:
        print("  Nenhuma entrada de pessoa fisica encontrada (todos sao CNPJ).")

    # 4. Coluna Municipio e codigo IBGE
    col_mun = _achar_coluna(df, "Município", "Municipio", "município", "municipio")
    df["_ibge6"] = df[col_mun].str.extract(r"^(\d{6})")
    df["_cidade"] = df[col_mun].str.extract(r"^\d+ - (.+)$")

    # 5. Filtrar municipios pequenos
    codigos_peq = {cod for cod, p in pop.items() if p <= LIMITE_POP}
    df_peq = df[df["_ibge6"].isin(codigos_peq)].copy()
    df_peq = df_peq.sort_values(["UF", "_cidade", "CNPJ Pesquisado"])
    print(f"  {len(df_peq)} registros em municipios ate {LIMITE_POP:,} habitantes.")

    if df_peq.empty:
        print("Nenhum registro para exportar.")
        return

    # 6. Selecionar colunas de saida (apenas as que existirem)
    colunas_presentes = [c for c in COLUNAS_SAIDA if c in df_peq.columns]
    # Garantir que Municipio e UF estejam presentes
    for c in (col_mun, "UF"):
        if c not in colunas_presentes:
            colunas_presentes.append(c)
    df_peq = df_peq[colunas_presentes]

    # 7. Lista de municipios ordenados
    municipios = (
        df_peq[["UF", col_mun]]
        .drop_duplicates()
        .sort_values(["UF", col_mun])
        .values.tolist()
    )
    print(f"  {len(municipios)} municipios pequenos com dados.")

    # 8. Gerar Excel com pandas
    print(f"\nGerando {ARQUIVO_SAIDA}...")
    with pd.ExcelWriter(ARQUIVO_SAIDA, engine="openpyxl") as writer:
        # Aba INDICE
        resumo = (
            df_peq.groupby(["UF", col_mun])
            .size()
            .reset_index(name="Qtd de CNPJs")
            .rename(columns={col_mun: "Município"})
        )
        resumo.to_excel(writer, sheet_name="ÍNDICE", index=False)

        for i, (uf, mun) in enumerate(municipios, 1):
            aba = sanitizar(f"{uf}-{mun}")
            df_mun = df_peq[(df_peq["UF"] == uf) & (df_peq[col_mun] == mun)]
            df_mun.to_excel(writer, sheet_name=aba, index=False)
            if i % 100 == 0:
                print(f"  {i}/{len(municipios)} abas gravadas...")

    # 9. Aplicar formatacao
    print("Aplicando formatacao...")
    wb = load_workbook(ARQUIVO_SAIDA)

    ws_idx = wb["ÍNDICE"]
    formatar_aba(ws_idx, len(resumo), len(resumo.columns))

    for i, (uf, mun) in enumerate(municipios, 1):
        aba = sanitizar(f"{uf}-{mun}")
        if aba not in wb.sheetnames:
            continue
        ws = wb[aba]
        n_linhas  = ws.max_row - 1
        n_colunas = ws.max_column
        if n_linhas > 0:
            formatar_aba(ws, n_linhas, n_colunas)
        if i % 100 == 0:
            print(f"  {i}/{len(municipios)} abas formatadas...")

    wb.save(ARQUIVO_SAIDA)

    print(f"\nConcluido!")
    print(f"  Arquivo  : {ARQUIVO_SAIDA}")
    print(f"  Municipios: {len(municipios)}")
    print(f"  CNPJs    : {len(df_peq)}")
    print(f"  Abas     : {len(municipios)} de municipios  +  1 INDICE")


if __name__ == "__main__":
    main()
