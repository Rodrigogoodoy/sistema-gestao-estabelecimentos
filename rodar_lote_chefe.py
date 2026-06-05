"""
Lote solicitado pelo chefe — PR + SC
Roda sequencialmente e salva um Excel por município.
"""
import datetime, time
from pathlib import Path

import pandas as pd
from playwright.sync_api import sync_playwright
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from bot_municipio import processar_municipio

MUNICIPIOS = [
    # PR
    {"uf": "PR", "municipio": "Lapa",               "cod_ibge": "4113205"},
    {"uf": "PR", "municipio": "Contenda",            "cod_ibge": "4106209"},
    {"uf": "PR", "municipio": "Campo Largo",         "cod_ibge": "4104204"},
    {"uf": "PR", "municipio": "Rio Negro",           "cod_ibge": "4122305"},
    {"uf": "PR", "municipio": "São Mateus do Sul",   "cod_ibge": "4125605"},
    {"uf": "PR", "municipio": "Três Barras do Paraná","cod_ibge": "4127858"},
    # SC (Rio Negrinho e São Bento do Sul são SC, não PR)
    {"uf": "SC", "municipio": "Rio Negrinho",        "cod_ibge": "4215000"},
    {"uf": "SC", "municipio": "São Bento do Sul",    "cod_ibge": "4215802"},
    {"uf": "SC", "municipio": "Fraiburgo",           "cod_ibge": "4205506"},
    {"uf": "SC", "municipio": "Videira",             "cod_ibge": "4219309"},
    {"uf": "SC", "municipio": "Caçador",             "cod_ibge": "4203006"},
    {"uf": "SC", "municipio": "Monte Carlo",         "cod_ibge": "4211058"},
    {"uf": "SC", "municipio": "Tangará",             "cod_ibge": "4217907"},
    {"uf": "SC", "municipio": "Curitibanos",         "cod_ibge": "4204806"},
]

URL_BASE = "https://cnes.datasus.gov.br/pages/estabelecimentos/consulta.jsp"

COLUNAS_MAP = {
    "CNPJ": "CNPJ", "CNPJ Mantenedora": "CNPJ Mantenedora",
    "Município": "Municipio", "UF": "UF", "CNES": "Codigo CNES",
    "Nome Fantasia": "Nome Fantasia", "Nome": "Razao Social",
    "Tipo de Estabelecimento": "Tipo de Estabelecimento",
    "Classificação Estabelecimento": "Classificacao", "Gestão": "Gestao",
    "Natureza Jurídica(Grupo)": "Natureza Juridica", "CEP": "CEP",
    "Logradouro": "Logradouro", "Número": "Numero", "Bairro": "Bairro",
    "Complemento": "Complemento", "Telefone": "Telefone", "E-mail": "Email",
    "Responsável Técnico": "Responsavel Tecnico",
    "Data Desativação": "Data Desativacao", "Motivo Desativação": "Motivo Desativacao",
}
LARGURAS = {
    "CNPJ": 20, "CNPJ Mantenedora": 22, "Municipio": 22, "UF": 8,
    "Codigo CNES": 14, "Nome Fantasia": 35, "Razao Social": 38,
    "Tipo de Estabelecimento": 30, "Classificacao": 30, "Gestao": 16,
    "Natureza Juridica": 30, "CEP": 14, "Logradouro": 32,
    "Numero": 10, "Bairro": 20, "Complemento": 20,
    "Telefone": 18, "Email": 28, "Responsavel Tecnico": 30,
    "Data Desativacao": 18, "Motivo Desativacao": 40,
}


def _limpar_cnpj(v):
    v = str(v or "").strip()
    return "" if v in ("---", "-", "nan", "None", "NaN") else v


def _fix_cnpj(row):
    cnpj    = _limpar_cnpj(row.get("CNPJ", ""))
    proprio = _limpar_cnpj(row.get("CNPJ Próprio", ""))
    mant    = _limpar_cnpj(row.get("CNPJ Mantenedora", ""))
    if cnpj and cnpj == mant and not proprio:
        cnpj = ""
    elif not cnpj and proprio:
        cnpj = proprio
    row["CNPJ"] = cnpj
    return row


def _gerar_excel(df, municipio, uf, output_path):
    cols   = {k: v for k, v in COLUNAS_MAP.items() if k in df.columns}
    df_exp = df[list(cols.keys())].rename(columns=cols)
    for col in ("Data Desativacao", "Motivo Desativacao"):
        if col not in df_exp.columns:
            df_exp[col] = ""
    df_exp = df_exp.replace({"null - null": "", "null": "", "---": ""}).fillna("")

    qtd      = len(df_exp)
    ativos   = int((df_exp["Data Desativacao"] == "").sum())
    inativos = qtd - ativos
    agora    = datetime.datetime.now().strftime("%d/%m/%Y")

    writer = pd.ExcelWriter(output_path, engine="openpyxl")
    df_exp.to_excel(writer, sheet_name="Estabelecimentos", index=False, startrow=4)
    ws = writer.sheets["Estabelecimentos"]

    def fill(cor): return PatternFill("solid", fgColor=cor)
    def borda():
        lado = Side(style="thin", color="D1D5DB")
        return Border(left=lado, right=lado, top=lado, bottom=lado)

    ncols = df_exp.shape[1]
    ult   = get_column_letter(ncols)

    ws.merge_cells(f"A1:{ult}1")
    ws["A1"] = f"ESTABELECIMENTOS DE SAUDE — {municipio.upper()}/{uf.upper()} — ATENDE SUS: SIM"
    ws["A1"].fill = fill("1E3A8A")
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells(f"A2:{ult}2")
    ws["A2"] = (f"Data de geracao: {agora}   |   Estabelecimentos: {qtd:,}"
                f"   |   Filtro: Atende SUS = Sim   |   Fonte: cnes.datasus.gov.br")
    ws["A2"].fill = fill("2563EB")
    ws["A2"].font = Font(bold=False, color="FFFFFF", size=10, name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    ws.merge_cells(f"A3:{ult}3")
    ws["A3"] = f"Ativos: {ativos:,}   |   Com data de desativação: {inativos:,}"
    ws["A3"].fill = fill("DBEAFE")
    ws["A3"].font = Font(bold=False, color="1E3A8A", size=10, name="Calibri")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 8

    for col in range(1, ncols + 1):
        c = ws.cell(row=5, column=col)
        c.fill = fill("1E3A8A")
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda()
    ws.row_dimensions[5].height = 30

    for row in range(6, qtd + 6):
        bg = "F0F9FF" if row % 2 == 0 else "FFFFFF"
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = fill(bg)
            c.font = Font(bold=False, color="1E293B", size=10, name="Calibri")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = borda()
        ws.row_dimensions[row].height = 18

    for i, col_name in enumerate(df_exp.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = LARGURAS.get(col_name, 18)

    ws.freeze_panes = "A6"
    writer.close()
    return qtd, ativos, inativos


def main():
    total = len(MUNICIPIOS)
    erros = []
    sucessos = []

    print(f"\n{'='*60}")
    print(f"  LOTE CHEFE — {total} municípios")
    print(f"{'='*60}\n")

    for i, item in enumerate(MUNICIPIOS, 1):
        uf  = item["uf"]
        mun = item["municipio"]
        cod = item["cod_ibge"]
        print(f"\n[{i}/{total}] {mun}/{uf}")
        print("-" * 40)
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                page    = browser.new_page()
                resultados = processar_municipio(
                    page=page, uf=uf, municipio=mun, cod_ibge=cod,
                    url_base=URL_BASE, delay_entre_fichas=0.3,
                )
                browser.close()

            if not resultados:
                print(f"  AVISO: nenhum estabelecimento encontrado.")
                erros.append(f"{mun}/{uf} — nenhum resultado")
                continue

            df = pd.DataFrame(resultados)
            df = df.apply(_fix_cnpj, axis=1)
            safe = mun.replace(" ", "_").replace("/", "_")
            arquivo = Path(f"resultado_{safe}_{uf}.xlsx")
            qtd, ativos, inativos = _gerar_excel(df, mun, uf, str(arquivo))
            print(f"  SALVO: {arquivo} — {qtd} registros ({ativos} ativos, {inativos} desativados)")
            sucessos.append(f"{mun}/{uf} → {arquivo} ({qtd} registros)")

        except Exception as e:
            print(f"  ERRO: {e}")
            erros.append(f"{mun}/{uf} — {e}")

    print(f"\n{'='*60}")
    print(f"  CONCLUÍDO: {len(sucessos)} OK, {len(erros)} erro(s)")
    print(f"{'='*60}")
    if sucessos:
        print("\nArquivos gerados:")
        for s in sucessos:
            print(f"  ✓ {s}")
    if erros:
        print("\nErros:")
        for e in erros:
            print(f"  ✗ {e}")
    print()


if __name__ == "__main__":
    main()
