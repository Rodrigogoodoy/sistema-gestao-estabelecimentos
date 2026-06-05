"""
Mescla os arquivos resultados_worker_N.xlsx e gera o relatorio final formatado.
Execute apos todos os workers terminarem: python merge_results.py
"""
import sys
import datetime
import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def merge():
    arquivos = sorted(Path(".").glob("resultados_worker_*.xlsx"))
    if not arquivos:
        print("Nenhum arquivo resultados_worker_*.xlsx encontrado.")
        sys.exit(1)

    print(f"Mesclando {len(arquivos)} arquivo(s)...\n")
    dfs = []
    for arq in arquivos:
        try:
            df = pd.read_excel(arq, dtype=str)
            dfs.append(df)
            print(f"  {arq.name}: {len(df):,} registros")
        except Exception as e:
            print(f"  {arq.name}: ERRO ao ler — {e}")

    if not dfs:
        print("\nNenhum dado para mesclar.")
        sys.exit(1)

    final = pd.concat(dfs, ignore_index=True)
    final.to_excel("resultados_final.xlsx", index=False)

    total = len(final)
    col   = final.get("ENCONTRADO", pd.Series(dtype=str))
    sim   = int((col == "SIM").sum())
    nao   = int((col == "NÃO").sum())
    erros = int((col == "ERRO").sum())

    print(f"\n{'='*48}")
    print(f"  RESULTADO FINAL: {total:,} registros")
    print(f"  SIM   : {sim:,}  ({sim/total*100:.1f}%)")
    print(f"  NAO   : {nao:,}  ({nao/total*100:.1f}%)")
    print(f"  ERRO  : {erros:,}  ({erros/total*100:.1f}%)")
    print(f"{'='*48}\n")

    print("Gerando ESTABELECIMENTOS_CNES.xlsx (apenas SIM)...")
    gerar_relatorio(final, total)
    print("\nPronto! Arquivos gerados:")
    print("  resultados_final.xlsx      — todos os resultados brutos")
    print("  ESTABELECIMENTOS_CNES.xlsx — somente os encontrados, formatado\n")


def gerar_relatorio(df: pd.DataFrame, total_pesquisados: int):
    sim_df = df[df["ENCONTRADO"] == "SIM"].copy().reset_index(drop=True)
    # Limpa valores inválidos
    for col in sim_df.columns:
        sim_df[col] = sim_df[col].replace({"null - null": "", "null": ""})
    qtd_sim = len(sim_df)
    pct_sim = round(qtd_sim / total_pesquisados * 100, 1) if total_pesquisados else 0
    agora   = datetime.datetime.now().strftime("%d/%m/%Y")

    colunas_map = {
        "CNPJ Pesquisado":               "CNPJ",
        "Município":                     "Municipio",
        "CNES":                          "Codigo CNES",
        "Nome Fantasia":                 "Nome Fantasia",
        "Nome":                          "Razao Social",
        "Tipo de Estabelecimento":       "Tipo de Estabelecimento",
        "Classificação Estabelecimento": "Classificacao",
        "Gestão":                        "Gestao",
        "Natureza Jurídica(Grupo)":      "Natureza Juridica",
        "UF":                            "UF",
        "CEP":                           "CEP",
        "Logradouro":                    "Logradouro",
        "Número":                        "Numero",
        "Bairro":                        "Bairro",
        "Telefone":                      "Telefone",
        "E-mail":                        "Email",
        "Responsável Técnico":           "Responsavel Tecnico",
        "Data Desativação":              "Data Desativacao",
        "Motivo Desativação":            "Motivo Desativacao",
    }
    colunas_existentes = {k: v for k, v in colunas_map.items() if k in sim_df.columns}
    sim_export = sim_df[list(colunas_existentes.keys())].rename(columns=colunas_existentes)

    writer = pd.ExcelWriter("ESTABELECIMENTOS_CNES.xlsx", engine="openpyxl")
    sim_export.to_excel(writer, sheet_name="Estabelecimentos", index=False, startrow=4)
    ws = writer.sheets["Estabelecimentos"]

    def fill(cor):   return PatternFill("solid", fgColor=cor)
    def borda():
        lado = Side(style="thin", color="D1D5DB")
        return Border(left=lado, right=lado, top=lado, bottom=lado)

    ncols     = sim_export.shape[1]
    ultima_col = get_column_letter(ncols)

    # Linha 1 — título
    ws.merge_cells(f"A1:{ultima_col}1")
    ws["A1"] = "ESTABELECIMENTOS DE SAUDE ENCONTRADOS NO CNES / DATASUS"
    ws["A1"].fill      = fill("1E3A8A")
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    # Linha 2 — subtítulo com stats
    ws.merge_cells(f"A2:{ultima_col}2")
    ws["A2"] = (
        f"Data de geracao: {agora}"
        f"   |   CNPJs pesquisados: {total_pesquisados:,}"
        f"   |   Encontrados: {qtd_sim:,} ({pct_sim}%)"
        f"   |   Fonte: cnes.datasus.gov.br"
    )
    ws["A2"].fill      = fill("2563EB")
    ws["A2"].font      = Font(bold=False, color="FFFFFF", size=10, name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # Linha 3 — nota
    ws.merge_cells(f"A3:{ultima_col}3")
    ws["A3"] = f"Lista com {qtd_sim:,} estabelecimentos confirmados no cadastro nacional de saude (CNES)"
    ws["A3"].fill      = fill("DBEAFE")
    ws["A3"].font      = Font(bold=False, color="1E3A8A", size=10, name="Calibri")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20

    # Linha 4 — espaço
    ws.row_dimensions[4].height = 8

    # Linha 5 — cabeçalho da tabela
    for col in range(1, ncols + 1):
        c = ws.cell(row=5, column=col)
        c.fill      = fill("1E3A8A")
        c.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = borda()
    ws.row_dimensions[5].height = 30

    # Linhas de dados
    for row in range(6, qtd_sim + 6):
        bg = "F0F9FF" if row % 2 == 0 else "FFFFFF"
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col)
            c.fill      = fill(bg)
            c.font      = Font(bold=False, color="1E293B", size=10, name="Calibri")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = borda()
        ws.row_dimensions[row].height = 18

    # Larguras
    larguras = {
        "CNPJ": 20, "Codigo CNES": 14, "Nome Fantasia": 35, "Razao Social": 38,
        "Tipo de Estabelecimento": 30, "Classificacao": 30, "Gestao": 16,
        "Natureza Juridica": 30, "Municipio": 22, "UF": 8, "CEP": 14,
        "Logradouro": 32, "Numero": 10, "Bairro": 20,
        "Telefone": 18, "Email": 28, "Responsavel Tecnico": 30,
        "Data Desativacao": 18, "Motivo Desativacao": 35,
    }
    for i, col_name in enumerate(sim_export.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(col_name, 20)

    ws.freeze_panes = "A6"
    writer.close()


if __name__ == "__main__":
    merge()
